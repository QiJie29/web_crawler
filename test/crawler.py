import requests
from bs4 import BeautifulSoup
import openpyxl
import re


def extract_data(input_string):
    pattern = r'\((.*?)\)'  # 正则表达式匹配括号内的内容
    match = re.search(pattern, input_string)

    if match:
        return match.group(1)
    else:
        return None


def getUrls():
    # 定义超链接内容
    base_url = "https://www.kylc.com/stats/global/yearly/"
    kind_url = "g_gdp/"
    year_min = 1960
    year_max = 2024
    suffix_url = ".html"

    # 定义urls数组用于存储超链接
    urls = list()

    # 循环遍历所有年份
    while year_min <= year_max:
        urls.append(base_url + kind_url + str(year_min) + suffix_url)
        year_min = year_min + 1
    return urls


def parser(urls):
    # 复制excel的sheet页面
    wb = openpyxl.load_workbook("../template.xlsx")
    # 提取模板sheet对象
    sheet_to_copy = wb["template data"]
    # 新建sheet
    new_sheet = wb.copy_worksheet(sheet_to_copy)
    # 重命名
    new_sheet.title = "result_data"
    # 定义初始化colIndex
    col_index = 3
    # 遍历urls数组
    for url in urls:
        r = requests.get(url)
        # 若html内容获取成功
        if r.status_code == 200:
            # 此处需使用decode解码，否则中文显示不正确
            soup = BeautifulSoup(r.content.decode("utf-8"), "html.parser")
            # print(soup)
            # 查询tbody标签中的区块
            table_node = soup.find("tbody")
            # print(table_node)
            # print("="*30)
            # 查询该区块中的td标签
            contents = table_node.find_all("td")
            # 定义数组索引号
            i = 0
            # 定义数据序列号
            count = 1
            # 提取当前页面年份作为列名
            col_index = col_index + 1
            title = soup.find("title").get_text()
            print(title[0:4])
            year = title[0:4]
            new_sheet.cell(1, col_index).value = year
            # print(len(contents))
            while i < len(contents):
                # print(contents[i].get_text())
                # 匹配当前标签内容为数据序列号，按照网页规则提取相应内容
                if contents[i].get_text() == str(count):
                    # print(contents[i].get_text() + " " + contents[i + 1].get_text() + " " + contents[i + 3].get_text())
                    # 写入excel中
                    row_min = 2
                    row_max = 216
                    # 循环匹配国家名称
                    while row_min <= row_max:
                        # 国家名称匹配成功后，将其数据填写至后方相应列中
                        # print(new_sheet['A' + str(row_min)].value+ contents[i + 1].get_text())
                        if new_sheet['A' + str(row_min)].value == contents[i + 1].get_text():
                            # 提取括号内数据
                            # print(extract_data(contents[i + 3].get_text()))
                            new_sheet.cell(row_min, col_index).value = extract_data(contents[i + 3].get_text())
                            break
                        else:
                            row_min = row_min + 1
                            # 用于输出未匹配成功的数据
                            if row_min > row_max:
                                print("未找到国家的名称:" + year + ":" + contents[i].get_text() + " " + contents[
                                    i + 1].get_text() + " " + contents[
                                          i + 3].get_text())
                    # 计数+1
                    count = count + 1
                    i = i + 5
                else:
                    i = i + 1
            # 测试环境下仅提取一份数据
            # break
        else:
            continue
    # 保存文件
    wb.save("result.xlsx")


if __name__ == '__main__':
    # 1.获取待爬取urls
    urls = list()
    urls = getUrls()

    # r = requests.get("https://www.kylc.com/stats/global/yearly/g_gdp/1960.html")
    # #解码
    # print(r.content.decode("utf-8"))

    # print(urls)
    # 2.解析urls.content
    parser(urls)

    pass
